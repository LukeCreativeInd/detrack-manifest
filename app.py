import streamlit as st
import pandas as pd
import openpyxl
import math
import zipfile
from io import BytesIO
import re
from datetime import datetime

st.set_page_config(page_title="CM Logistics Manifest Generator", layout="centered")

# Display logo
st.image("CM_Logistics_Top_Logo.png", use_container_width=True)

st.title("CM Logistics Manifest Generator")
st.markdown("Upload your orders export CSV, choose the group, then click Generate to get your manifests.")

# Group selector
group_option = st.selectbox("Select Group Name:", ["Clean Eats Australia", "Made Active"])

# Cold Express checkbox
cold_required = st.checkbox("Is there a Cold Express Pickup Required?")

uploaded_file = st.file_uploader("Upload orders_export CSV file", type="csv")

generate = st.button("Generate Manifests")

if uploaded_file and generate:
    orders_df = pd.read_csv(uploaded_file)

    # Clean and prep
    orders_df.columns = orders_df.columns.str.strip()
    orders_df["Notes"] = orders_df["Notes"].fillna("")
    orders_df["Tags"] = orders_df["Tags"].fillna("")

    def format_phone(phone):
        if pd.isna(phone):
            return ""
        phone = str(phone).strip().replace(" ", "").replace("+", "")
        if phone.startswith("61"):
            phone = "0" + phone[2:]
        elif phone.startswith("4"):
            phone = "0" + phone
        return phone

    state_map = {"VIC": "Victoria", "NSW": "New South Wales"}
    country_map = {"AU": "Australia"}

    manifest_rows = []
    grouped_orders = orders_df.groupby("Name", sort=False)  # 'Name' is the order number

    for name, group in grouped_orders:
        order = group.iloc[0]
        order_number = name
        bundle_items = [
            "CARB LOVER'S FEAST",
            "SUPER CHARGED CALORIES",
            "FEED ME BEEF",
            "GIVE ME CHICKEN",
            "I WON'T PAS(TA) ON THIS MEAL",
            "THE MEGA PACK",
            "MAKE YOUR OWN MEGA PACK",
            "CARB HATERS FEAST",
            "UNDER CHARGED CALORIES",
            "VEGGIE LOVERS PACK",
            "Clean Eats Meal Plan"
        ]

        made_active_bundles = {
            "10 Pack": 10,
            "20 Pack": 20,
            "30 Pack": 30,
            "10 Meal Christmas Bundle": 10,
            "14 Meal Christmas Bundle": 14,
            "High Protein Pack": 12,
            "The Bunny Bundle": 10
        }
        if group_option == "Clean Eats Australia":
            group["Lineitem name"] = group["Lineitem name"].astype(str).str.strip()
            non_bundle_items = group[~group["Lineitem name"].isin(bundle_items)]
            total_qty = non_bundle_items["Lineitem quantity"].sum()
        else:
            total_qty = 0
            for _, row in group.iterrows():
                name = row["Lineitem name"]
                qty = row["Lineitem quantity"]
                if name in made_active_bundles:
                    total_qty += made_active_bundles[name] * qty
                else:
                    total_qty += qty
        labels = math.ceil(total_qty / 20)

        phone = order.get("Billing Phone", "")
        if pd.isna(phone) or phone == "":
            phone = order.get("Phone", "")
        phone = format_phone(phone)

        state = state_map.get(order["Shipping Province"], order["Shipping Province"])
        country = country_map.get(order["Shipping Country"], order["Shipping Country"])
        city = "Melbourne" if state == "Victoria" else "Sydney" if state == "New South Wales" else ""

        date_match = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", order["Tags"])
        delivery_date = date_match.group(1) if date_match else ""

        manifest_rows.append({
            "D.O. No.": order_number,
            "Date": delivery_date,
            "Address 1": order["Shipping Street"],
            "Address 2": order["Shipping City"],
            "Postal Code": str(order["Shipping Zip"]).replace("'", ""),
            "State": state,
            "Country": country,
            "Deliver to": order["Shipping Name"],
            "Phone No.": phone,
            "Time Window": "0600-1800",
            "City": city,
            "Group": group_option,
            "No. of Shipping Labels": labels,
            "Line Items": total_qty,
            "Instructions": order["Notes"]
        })

    manifest_df = pd.DataFrame(manifest_rows)

    cm_names = orders_df[orders_df["Tags"].str.contains("CM")]["Name"].unique()
    mc_names = orders_df[orders_df["Tags"].str.contains("MC")]["Name"].unique()
    cx_names = orders_df[orders_df["Tags"].str.contains("CX")]["Name"].unique()
    all_tagged_names = set(cm_names) | set(mc_names) | set(cx_names)

    cm_manifest = manifest_df[manifest_df["D.O. No."].isin(cm_names)]
    mc_manifest = manifest_df[manifest_df["D.O. No."].isin(mc_names)]
    cx_manifest = manifest_df[manifest_df["D.O. No."].isin(cx_names)]
    other_manifest = manifest_df[~manifest_df["D.O. No."].isin(all_tagged_names)]

    # Add Cold Express row if selected
    if cold_required:
        total_cartons = int(cx_manifest["No. of Shipping Labels"].sum()) if not cx_manifest.empty else ""
        today_str = datetime.now().strftime("%d/%m/%Y")
        cold_row = {
            "D.O. No.": "CXMANIFEST",
            "Date": today_str,
            "Address 1": "830 Wellington Rd",
            "Address 2": "Rowville",
            "Postal Code": 3178,
            "State": "Victoria",
            "Country": "Australia",
            "Deliver to": "Cold Xpress",
            "Phone No.": "",
            "Time Window": "0600-1800",
            "City": "Melbourne",
            "Group": "Clean Eats Australia",
            "No. of Shipping Labels": total_cartons,
            "Line Items": "",
            "Instructions": ""
        }
        mc_manifest = pd.concat([mc_manifest, pd.DataFrame([cold_row])], ignore_index=True)

    # Build CX Ready Manifest
    output = BytesIO()
    with zipfile.ZipFile(output, "w") as zipf:
        def add_to_zip(df, filename):
            if df.empty:
                return
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                if "Phone No." in df.columns:
                    df["Phone No."] = df["Phone No."].astype(str).str.replace(r"\.0$", "", regex=True)
                df.to_excel(writer, index=False, sheet_name='Manifest')
                workbook = writer.book
                worksheet = writer.sheets['Manifest']
                text_fmt = workbook.add_format({'num_format': '@'})
                if "Phone No." in df.columns:
                    col_index = df.columns.get_loc("Phone No.")
                    worksheet.set_column(col_index, col_index, None, text_fmt)
            zipf.writestr(filename, buffer.getvalue())

        add_to_zip(cm_manifest, "CM_Manifest.xlsx")
        add_to_zip(mc_manifest, "MC_Manifest.xlsx")
        add_to_zip(cx_manifest, "CX_Manifest.xlsx")
        add_to_zip(other_manifest, "Other_Manifest.xlsx")

        if not cx_manifest.empty:
            from datetime import timedelta
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from tempfile import NamedTemporaryFile

            cx_ready_body = pd.DataFrame({
                "INV NO.": cx_manifest["D.O. No."],
                "DELIVERY DATE": pd.to_datetime(cx_manifest["Date"], format="%d/%m/%Y", errors='coerce') + timedelta(days=1),
                "STORE NO": "",
                "STORE NAME": cx_manifest["Deliver to"],
                "ADDRESS": cx_manifest["Address 1"],
                "SUBURB": cx_manifest["Address 2"],
                "STATE": cx_manifest["State"],
                "POSTCODE": cx_manifest["Postal Code"],
                "CARTONS": cx_manifest["No. of Shipping Labels"],
                "PALLETS": "",
                "WEIGHT (KG)": (cx_manifest["Line Items"].astype(float) * 0.4).round(2),
                "INV. VALUE": "",
                "COD": "",
                "TEMP": "chilled",
                "COMMENT": cx_manifest["Instructions"]
            })
            cx_ready_body["DELIVERY DATE"] = cx_ready_body["DELIVERY DATE"].dt.strftime("%d/%m/%Y")

            wb = load_workbook("cx_manifest_template.xlsx")
            ws = wb.active

            for r_idx, row in enumerate(dataframe_to_rows(cx_ready_body, index=False, header=False), start=6):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    if cell.coordinate in ws.merged_cells:
                        continue
                    safe_value = "" if pd.isna(value) else str(value)
                    cell.value = safe_value

            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                zipf.writestr("CX_Ready_Manifest.xlsx", tmp.read())

    output.seek(0)
    st.download_button(
        label="Download Manifests ZIP",
        data=output,
        file_name="Meal_Cart_Manifests.zip",
        mime="application/zip"
    )
