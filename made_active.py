import streamlit as st
import pandas as pd
import math
import zipfile
from io import BytesIO
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from tempfile import NamedTemporaryFile

def run():
    st.markdown("### Made Active Manifest Generator")

    uploaded_file = st.file_uploader("Upload Made Active orders_export CSV file", type="csv")

    if not uploaded_file:
        return

    orders_df = pd.read_csv(uploaded_file)
    orders_df.columns = orders_df.columns.str.strip()
    orders_df["Notes"] = orders_df["Notes"].fillna("")
    orders_df["Tags"] = orders_df["Tags"].fillna("")

    bundle_map = {
        "10 Pack": 10,
        "20 Pack": 20,
        "30 Pack": 30,
        "10 Meal Christmas Bundle": 10,
        "14 Meal Christmas Bundle": 14,
        "High Protein Pack": 12,
        "The Bunny Bundle": 10
    }

    def format_phone(phone):
        if pd.isna(phone):
            return ""
        phone = str(phone).strip().replace(" ", "").replace("+", "")
        if phone.startswith("61"):
            phone = "0" + phone[2:]
        elif phone.startswith("4"):
            phone = "0" + phone
        return phone

    manifest_rows = []
    grouped_orders = orders_df.groupby("Name", sort=False)

    for name, group in grouped_orders:
        order = group.iloc[0]
        total_qty = 0
        for _, row in group.iterrows():
            item = row["Lineitem name"].strip()
            qty = row["Lineitem quantity"]
            if item in bundle_map:
                total_qty += bundle_map[item] * qty
            else:
                total_qty += qty

        labels = math.ceil(total_qty / 20)
        phone = order.get("Billing Phone") or order.get("Phone")
        phone = format_phone(phone)

        state_map = {
            "VIC": "Victoria",
            "NSW": "New South Wales",
            "ACT": "Australian Capital Territory"
        }
        country_map = {"AU": "Australia"}
        state = state_map.get(order["Shipping Province"], order["Shipping Province"])
        country = country_map.get(order["Shipping Country"], order["Shipping Country"])

        date_match = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", order["Tags"])
        delivery_date = date_match.group(1) if date_match else ""

        manifest_rows.append({
            "D.O. No.": name,
            "Date": delivery_date,
            "Address 1": order["Shipping Street"],
            "Address 2": order["Shipping City"],
            "Postal Code": str(order["Shipping Zip"]).replace("'", ""),
            "State": state,
            "Country": country,
            "Deliver to": order["Shipping Name"],
            "Phone No.": phone,
            "Time Window": "0600-1800",
            "Group": "Made Active",
            "No. of Shipping Labels": labels,
            "Line Items": total_qty,
            "Email": order["Email"],
            "Instructions": order["Notes"]
        })

    manifest_df = pd.DataFrame(manifest_rows)

    cm_names = orders_df[orders_df["Tags"].str.contains("CM")]["Name"].unique()
    mc_names = orders_df[orders_df["Tags"].str.contains("MC")]["Name"].unique()
    cx_names = orders_df[orders_df["Tags"].str.contains("CX")]["Name"].unique()
    all_tagged_names = set(cm_names) | set(mc_names) | set(cx_names)

    cm_manifest = manifest_df[manifest_df["D.O. No."].isin(cm_names)]
    mc_manifest = manifest_df[manifest_df["D.O. No."].isin(mc_names)]
    cx_manifest = manifest_df[manifest_df["D.O. No."].isin(cx_names)]
    other_manifest = manifest_df[~manifest_df["D.O. No."].isin(all_tagged_names)]

    # --- Polar Parcel Manifest Logic ---
    polar_states = ["New South Wales", "Australian Capital Territory"]
    polar_df = manifest_df[manifest_df["State"].isin(polar_states)]

    output = BytesIO()
    with zipfile.ZipFile(output, "w") as zipf:
        def add_to_zip(df, filename):
            if df.empty:
                return
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                df["Phone No."] = df["Phone No."].astype(str).str.replace(r"\.0$", "", regex=True)
                df.to_excel(writer, index=False, sheet_name='Manifest')
                workbook = writer.book
                worksheet = writer.sheets['Manifest']
                text_fmt = workbook.add_format({'num_format': '@'})
                col_index = df.columns.get_loc("Phone No.")
                worksheet.set_column(col_index, col_index, None, text_fmt)
            zipf.writestr(filename, buffer.getvalue())

        add_to_zip(cm_manifest, "CM_Manifest.xlsx")
        add_to_zip(mc_manifest, "MC_Manifest.xlsx")
        add_to_zip(cx_manifest, "CX_Manifest.xlsx")
        add_to_zip(other_manifest, "Other_Manifest.xlsx")

        # --- Polar Parcel Manifest Generation ---
        if not polar_df.empty:
            polar_orders = []
            for _, row in polar_df.iterrows():
                polar_orders.append([
                    "Made Active",             # Seller Name
                    row["D.O. No."],          # Order No.
                    row["Deliver to"],        # Customer Name
                    row["Address 1"],         # Address
                    row["Address 2"],         # City
                    row["Postal Code"],       # Postcode
                    str(row["Phone No."]) if row["Phone No."] else "",  # Phone as text
                    row["Email"],             # Email
                    row["Instructions"],      # Delivery Notes
                    row["No. of Shipping Labels"]  # Cartons
                ])

            wb = Workbook()
            ws = wb.active

            # Row 1: "Delivery Date" (A1), <date + 2 days> (B1)
            ws["A1"] = "Delivery Date"
            delivery_date = (datetime.now() + timedelta(days=2)).strftime("%d/%m/%Y")
            ws["B1"] = delivery_date

            # Row 2: Headers (A2:J2)
            headers = [
                "Seller Name", "Order No.", "Customer Name", "Address", "City",
                "Postcode", "Phone", "Email", "Delivery Notes", "Cartons"
            ]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=2, column=col_num, value=header)

            # Order data (Row 3 down), with phone as text
            for row_idx, order in enumerate(polar_orders, start=3):
                for col_idx, value in enumerate(order, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 7:  # Phone column
                        cell.number_format = '@'

            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                zipf.writestr("Polar_Parcel_Manifest.xlsx", tmp.read())

    output.seek(0)
    st.download_button(
        label="Download Manifests ZIP",
        data=output,
        file_name="MadeActive_Manifests.zip",
        mime="application/zip"
    )
